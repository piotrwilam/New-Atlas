"""Read the per-concept neuron-list XLSX artifacts produced by the
decomposition stage.

Each XLSX has columns: object, layer, n_concept_only, n_both,
n_token_only, concept_only, both, token_only — where the three "_only"
list columns are stored as string-encoded Python lists of int neuron IDs.

Files are split into two layer halves (part1, part2). This module
transparently concatenates both halves.
"""

from __future__ import annotations

import ast
from pathlib import Path

import openpyxl

from atlas.paths import DATA_ROOT

_VALID_MODELS = {"QW", "DS"}
_VALID_LANGS = {"P", "R"}
# Python data uses two prefixes — `ast__` for syntax-tree concepts (Import,
# While, For, …) and `builtin__` for callable builtins (len, range, sorted,
# …). Both are stripped on load so the dendrogram leaf labels match the
# paper figures (which show just `len`, not `builtin__len`).
_PREFIXES: dict[str, tuple[str, ...]] = {
    "P": ("ast__", "builtin__"),
    "R": ("rust__",),
}


def _strip_prefix(name: str, prefixes: tuple[str, ...]) -> str:
    for p in prefixes:
        if name.startswith(p):
            return name[len(p):]
    return name


def load_neuron_lists(
    *,
    model: str,
    lang: str,
    eps: float,
    cons: float,
    layer: int,
    partition: str = "concept_only",
    data_root: Path | None = None,
) -> dict[str, set[int]]:
    """Load concept-only / shared / token-only neuron sets at one (model,
    lang, layer) cell from the decomposition-stage XLSX artifacts.

    Parameters
    ----------
    model      one of {"QW", "DS"} — Qwen or DeepSeek
    lang       one of {"P", "R"}   — Python or Rust
    eps        epsilon threshold used for binarisation (e.g. 0.5)
    cons       consistency threshold (e.g. 0.8)
    layer      integer layer index (0-based)
    partition  which column to return: "concept_only", "both", or "token_only"
    data_root  override the global DATA_ROOT (e.g. for tests)

    Returns
    -------
    dict mapping concept name (with the language prefix stripped — e.g.
    "Import" not "ast__Import") to a set of int neuron IDs at the given
    layer for the given partition.

    Side effects: reads two XLSX files from disk.
    """
    assert model in _VALID_MODELS, f"model must be one of {_VALID_MODELS}, got {model!r}"
    assert lang in _VALID_LANGS, f"lang must be one of {_VALID_LANGS}, got {lang!r}"
    assert partition in {"concept_only", "both", "token_only"}, (
        f"partition must be concept_only|both|token_only, got {partition!r}"
    )
    assert isinstance(layer, int) and layer >= 0, f"layer must be non-negative int, got {layer!r}"

    root = Path(data_root) if data_root is not None else DATA_ROOT
    prefixes = _PREFIXES[lang]
    col_index = {"concept_only": 5, "both": 6, "token_only": 7}[partition]

    out: dict[str, set[int]] = {}
    for part in (1, 2):
        fname = f"{lang}_{model}_4_neuron_list_eps{eps}_cons{cons}_layers_part{part}_both.xlsx"
        path = root / fname
        if not path.exists():
            raise FileNotFoundError(f"Expected neuron-list file not found: {path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if row[1] != layer:
                continue
            obj = row[0]
            cell = row[col_index]
            try:
                ids = set(ast.literal_eval(cell)) if cell else set()
            except (ValueError, SyntaxError) as e:
                raise ValueError(
                    f"Could not parse neuron list for {obj!r} L{layer} in {path}: {cell!r}"
                ) from e
            out[_strip_prefix(obj, prefixes)] = ids
    return out


def load_concept_sizes_by_layer(
    *,
    model: str,
    lang: str,
    eps: float,
    cons: float,
    partition: str = "concept_only",
    data_root: Path | None = None,
) -> dict[str, dict[int, int]]:
    """Load |partition set| per concept per layer.

    Reads the n_{partition} count column directly — much faster than
    `load_neuron_lists` because it skips the string-parsed neuron list.
    Used by per-layer temporal-dynamics figures (F4) and circuit-size
    profiles (F9–F12).

    The `"universal"` partition is the universal mask A = concept_only ∪
    both (i.e. every neuron in the universal circuit, regardless of
    whether the keyword token also fires). This is the "Circuit size"
    measure used in F4 and the §5.1 mean-concept-fraction table.

    Returns
    -------
    dict mapping concept name (prefix stripped) → dict mapping layer → size.
    """
    assert model in _VALID_MODELS, f"model must be one of {_VALID_MODELS}, got {model!r}"
    assert lang in _VALID_LANGS, f"lang must be one of {_VALID_LANGS}, got {lang!r}"
    valid_partitions = {"concept_only", "both", "token_only", "universal"}
    assert partition in valid_partitions, (
        f"partition must be one of {sorted(valid_partitions)}, got {partition!r}"
    )

    root = Path(data_root) if data_root is not None else DATA_ROOT
    prefixes = _PREFIXES[lang]

    out: dict[str, dict[int, int]] = {}
    for part in (1, 2):
        fname = f"{lang}_{model}_4_neuron_list_eps{eps}_cons{cons}_layers_part{part}_both.xlsx"
        path = root / fname
        if not path.exists():
            raise FileNotFoundError(f"Expected neuron-list file not found: {path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            obj = row[0]
            layer = row[1]
            if obj is None or layer is None:
                continue
            if partition == "universal":
                size = int(row[2] or 0) + int(row[3] or 0)
            else:
                col = {"concept_only": 2, "both": 3, "token_only": 4}[partition]
                size = int(row[col] or 0)
            name = _strip_prefix(obj, prefixes)
            out.setdefault(name, {})[int(layer)] = size
    return out


def load_flow_type_assignments(
    *,
    model: str,
    lang: str,
    data_root: Path | None = None,
) -> dict[str, str]:
    """Per-concept flow-type assignments at one (lang, model) cell.

    Flow type is the §6.1 classification of each concept's circuit-size-by-
    layer profile into one of: `two_phase`, `build_and_hold`,
    `late_emergence`, `unclassified`. Used by paper figures F9-F12 to
    colour-code per-concept circuit-size curves.

    Reads `7_E6_flow_type_assignments.xlsx` from DATA_ROOT — a single file
    that holds all four (lang, model) cells; this function filters to one.

    Returns
    -------
    dict mapping concept name (prefix stripped) → flow_type string.
    """
    assert model in _VALID_MODELS, f"model must be one of {_VALID_MODELS}, got {model!r}"
    assert lang in _VALID_LANGS, f"lang must be one of {_VALID_LANGS}, got {lang!r}"

    root = Path(data_root) if data_root is not None else DATA_ROOT
    path = root / "7_E6_flow_type_assignments.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Expected flow-type file not found: {path}")

    prefixes = _PREFIXES[lang]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        row_lang, row_model, concept, flow_type = row[0], row[1], row[2], row[3]
        if row_lang != lang or row_model != model:
            continue
        if concept is None or flow_type is None:
            continue
        out[_strip_prefix(concept, prefixes)] = flow_type
    return out
